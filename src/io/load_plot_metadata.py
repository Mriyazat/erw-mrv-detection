"""Per-plot Metadata-sheet extractor (GPS, serial, firmware, n_configs).

Mines the `Metadata` sheet from all 24 plot files into a tidy parquet for
the geostatistics work (per-plot lat/lon) and full sensor traceability.
"""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl
import pandas as pd

from src.config import CACHE_DIR, PLOT_FILES_2025, PLOT_FILES_2026, TREATMENT_MAP

logger = logging.getLogger(__name__)

CACHE_PATH = CACHE_DIR / "plot_metadata.parquet"

KEY_FIELDS = {
    "Device Name": "device_name", "Serial Number": "serial_number",
    "Device Type": "device_type", "Firmware Version": "firmware_version",
    "Hardware Version": "hardware_version",
    "Measurement Interval": "measurement_interval",
    "Software Version": "software_version", "Latitude": "latitude",
    "Longitude": "longitude", "Logger Time": "logger_time",
    "Time Zone": "time_zone", "Satellite Vehicles": "gps_satellites",
    "GPS Fix Status": "gps_fix_status",
    "Horizontal Accuracy": "gps_horizontal_accuracy", "Altitude": "altitude_m",
    "Modem Firmware Version": "modem_firmware", "Modem Type": "modem_type",
}


def _parse_metadata_sheet(filepath: Path) -> dict:
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if "Metadata" not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb["Metadata"]
    record: dict[str, object] = {}
    for r in range(1, ws.max_row + 1):
        k = ws.cell(r, 2).value
        v = ws.cell(r, 3).value
        if isinstance(k, str) and k.strip() in KEY_FIELDS:
            record[KEY_FIELDS[k.strip()]] = v
    record["n_configs"] = sum(
        1 for s in wb.sheetnames if s.startswith("Processed Data Config"))
    record["sheet_count"] = len(wb.sheetnames)
    wb.close()
    return record


def _coerce(record: dict) -> dict:
    for col in ("latitude", "longitude", "altitude_m"):
        v = record.get(col)
        if v is None:
            continue
        try:
            record[col] = float(v)
        except (TypeError, ValueError):
            record[col] = None
    for col in ("gps_satellites", "gps_fix_status", "gps_horizontal_accuracy",
                "modem_firmware", "modem_type", "n_configs", "sheet_count",
                "hardware_version"):
        v = record.get(col)
        if v is None:
            continue
        try:
            record[col] = int(v)
        except (TypeError, ValueError):
            try:
                record[col] = float(v)
            except (TypeError, ValueError):
                pass
    return record


def load_plot_metadata(use_cache: bool = True) -> pd.DataFrame:
    if use_cache and CACHE_PATH.exists():
        logger.info("Loading cached plot metadata from %s", CACHE_PATH)
        return pd.read_parquet(CACHE_PATH)

    rows = []
    for deployment, file_map in (("2025_oct", PLOT_FILES_2025),
                                  ("2026_apr", PLOT_FILES_2026)):
        for plot_id, fp in file_map.items():
            fp = Path(fp)
            if not fp.exists():
                logger.warning("Missing plot file: %s", fp)
                continue
            logger.info("Mining metadata from %s", fp.name)
            rec = _coerce(_parse_metadata_sheet(fp))
            rec["plot_id"] = plot_id
            rec["treatment"] = TREATMENT_MAP.get(plot_id, "unknown")
            rec["deployment"] = deployment
            rec["source_file"] = fp.name
            rows.append(rec)

    df = pd.DataFrame(rows)
    front = ["plot_id", "treatment", "deployment", "device_name",
             "serial_number", "latitude", "longitude", "altitude_m",
             "gps_fix_status", "gps_satellites", "gps_horizontal_accuracy",
             "firmware_version", "hardware_version", "n_configs",
             "logger_time", "time_zone", "source_file"]
    front = [c for c in front if c in df.columns]
    rest = [c for c in df.columns if c not in front]
    df = (df[front + rest].sort_values(["plot_id", "deployment"])
            .reset_index(drop=True))

    CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    df.to_parquet(CACHE_PATH, index=False)
    logger.info("Cached %d plot-metadata rows to %s", len(df), CACHE_PATH)
    return df


def per_plot_centroid(meta: pd.DataFrame) -> pd.DataFrame:
    return (meta.dropna(subset=["latitude", "longitude"])
              .groupby("plot_id")
              .agg(latitude=("latitude", "mean"),
                   longitude=("longitude", "mean"),
                   altitude_m=("altitude_m", "mean"),
                   gps_horizontal_accuracy=("gps_horizontal_accuracy", "mean"),
                   n_deployments=("deployment", "nunique"),
                   max_n_configs=("n_configs", "max"),
                   treatment=("treatment", "first"),
                   serial_number=("serial_number", "first"))
              .reset_index())


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s %(levelname)s %(message)s")
    df = load_plot_metadata(use_cache=False)
    print(f"\nShape: {df.shape}")
    print(per_plot_centroid(df).to_string(index=False))
