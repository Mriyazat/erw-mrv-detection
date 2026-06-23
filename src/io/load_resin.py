"""UNIBEST PRS resin ion-flux capsule loader.

Parses both UNIBEST International workbooks into a single tidy parquet:
  data/resin/resin_R1R2.xlsx  (Round 1 + Round 2)
  data/resin/resin_R3.xlsx    (Round 3 only)

Sample ID grammar ::= "<plot> <half> <depth_label> [Round <round_num>]"
  plot in {1..9}; half in {W,E}; depth_label in {Shallow,Middle,Deep};
  round_num in {1,2,3} (absent in the R3-only workbook).
"""

from __future__ import annotations

import logging
import re
from pathlib import Path

import openpyxl
import pandas as pd

from src.config import (
    CACHE_DIR,
    RESIN_ANALYTE_COLS,
    RESIN_DEPTH_MAP,
    RESIN_FILES,
    RESIN_QA_FLAGS,
    RESIN_ROUND_DATES,
    TREATMENT_MAP,
)

logger = logging.getLogger(__name__)

CACHE_PATH = CACHE_DIR / "resin.parquet"

SAMPLE_RE = re.compile(
    r"^\s*(?P<plot>\d+)\s+(?P<half>[WE])\s+(?P<depth>Shallow|Middle|Deep)"
    r"(?:\s+Round\s+(?P<round>\d+))?\s*$",
    re.IGNORECASE,
)

HEADER_KEYS = ("Barcode", "Sample ID", "Depth Low", "Depth High")


def _find_data_start(ws) -> int | None:
    for r in range(1, ws.max_row + 1):
        first = ws.cell(r, 1).value
        if isinstance(first, str) and first.strip().lower() == "barcode":
            return r
    return None


def _parse_sample_id(sample_id: str, default_round: int | None) -> dict | None:
    if not isinstance(sample_id, str):
        return None
    m = SAMPLE_RE.match(sample_id)
    if not m:
        return None

    plot = int(m.group("plot"))
    half = m.group("half").upper()
    depth_label = m.group("depth").capitalize()
    rn = m.group("round")
    if rn is None:
        if default_round is None:
            logger.warning("Sample %r has no Round and no default", sample_id)
            return None
        rn = default_round
    else:
        rn = int(rn)

    plot_half = f"{plot}{half}"
    return {
        "plot_id": plot,
        "half": half,
        "plot_half": plot_half,
        "depth_label": depth_label,
        "depth_cm": RESIN_DEPTH_MAP[depth_label],
        "round": rn,
        "treatment": TREATMENT_MAP.get(plot_half, "unknown"),
    }


def _load_workbook_table(path: Path, default_round: int | None) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(path)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet 1"]

    header_row = _find_data_start(ws)
    if header_row is None:
        raise ValueError(f"No 'Barcode' header row found in {path.name}")

    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    headers_str = [str(h).strip() if h is not None else "" for h in headers]
    col_idx = {h: i for i, h in enumerate(headers_str) if h}

    missing = [k for k in HEADER_KEYS if not any(h.startswith(k) for h in col_idx)]
    if missing:
        raise ValueError(f"{path.name}: missing required header columns {missing}")

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        raw = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if all(v is None for v in raw):
            continue
        barcode_val = raw[col_idx.get("Barcode", 0)]
        sample_id = raw[col_idx.get("Sample ID", 1)]
        if barcode_val is None and sample_id is None:
            continue

        parsed = _parse_sample_id(
            str(sample_id) if sample_id is not None else "", default_round,
        )
        if parsed is None:
            logger.warning("%s row %d: cannot parse Sample ID %r - skipping",
                           path.name, r, sample_id)
            continue

        rec = {
            "barcode": str(barcode_val) if barcode_val is not None else "",
            **parsed,
            "source_file": path.name,
        }
        for analyte_name, out_col in RESIN_ANALYTE_COLS.items():
            val = raw[col_idx[analyte_name]] if analyte_name in col_idx else None
            try:
                rec[out_col] = float(val) if val is not None and val != "" else None
            except (TypeError, ValueError):
                rec[out_col] = None

        rec["qa_flag"] = RESIN_QA_FLAGS.get(rec["barcode"], "")
        rows.append(rec)

    wb.close()
    return pd.DataFrame(rows)


def load_resin(use_cache: bool = True) -> pd.DataFrame:
    """Return the cleaned 121-row resin table."""
    if use_cache and CACHE_PATH.exists():
        logger.info("Loading cached resin data from %s", CACHE_PATH)
        return pd.read_parquet(CACHE_PATH)

    file_round_default = {"R1_R2": None, "R3": 3}

    frames = []
    for tag, path in RESIN_FILES.items():
        logger.info("Reading resin workbook %s (default_round=%s)",
                    path.name, file_round_default[tag])
        df = _load_workbook_table(path, default_round=file_round_default[tag])
        logger.info("  -> %d rows from %s", len(df), path.name)
        frames.append(df)

    resin = pd.concat(frames, ignore_index=True)

    deploy = pd.DataFrame(
        [(r, *RESIN_ROUND_DATES[r]) for r in RESIN_ROUND_DATES],
        columns=["round", "deploy_start", "deploy_end"],
    )
    deploy["deploy_start"] = pd.to_datetime(deploy["deploy_start"])
    deploy["deploy_end"]   = pd.to_datetime(deploy["deploy_end"])
    deploy["days_deployed"] = (
        (deploy["deploy_end"] - deploy["deploy_start"]).dt.total_seconds() / 86400.0
    )
    resin = resin.merge(deploy, on="round", how="left")

    resin["round"]    = resin["round"].astype("int8")
    resin["plot_id"]  = resin["plot_id"].astype("int8")
    resin["depth_cm"] = resin["depth_cm"].astype("int16")

    front = [
        "barcode", "plot_id", "half", "plot_half", "depth_cm", "depth_label",
        "round", "deploy_start", "deploy_end", "days_deployed", "treatment",
    ]
    analyte_cols = list(RESIN_ANALYTE_COLS.values())
    tail = ["qa_flag", "source_file"]
    resin = resin[front + analyte_cols + tail]

    CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    resin.to_parquet(CACHE_PATH, index=False)
    logger.info("Cached %d resin rows to %s", len(resin), CACHE_PATH)
    return resin


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s %(levelname)s %(message)s")
    df = load_resin(use_cache=False)
    print(f"\nShape: {df.shape}\n")
    print(df.groupby(["round", "depth_cm"]).size().unstack(fill_value=0))
    print(f"\nQA flags set: {(df['qa_flag'] != '').sum()}")
