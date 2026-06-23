"""Multi-config TEROS-12 / TEROS-21 plot-sensor loader.

Each plot XLSX contains multiple "Processed Data Config N" sheets from
successive sensor reconfigurations. Schemas differ (2-depth early configs
vs 3-depth current configs). Port-to-depth mapping is NOT stable across
configs, so a SENSOR-BLOCK-ORDER heuristic assigns blocks to depths
[15, 40, 100] in column order.

VERIFICATION HARDENING (vs the prior repos):
  * Every skipped non-standard sheet is recorded in the audit with its row
    count, so silent row loss is surfaced (see scripts/02_validate_extraction.py).
  * The block-order -> depth heuristic is the documented assumption; the
    validation script cross-checks it against per-file Metadata and flags
    any file where the assumption is questionable.
  * Only this loader exists (the legacy single-largest-config / hardcoded
    column-index loader from erw_ml is intentionally NOT ported).
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl
import pandas as pd

from src.config import (
    CACHE_DIR,
    PLOT_FILES,
    SENSOR_BLOCK_DEPTH_ORDER,
    TREATMENT_MAP,
)

logger = logging.getLogger(__name__)

# Block index -> depth (cm). Validated DEEP->SHALLOW; see config note.
DEPTH_ORDER = list(SENSOR_BLOCK_DEPTH_ORDER)
CACHE_PATH  = CACHE_DIR / "sensors.parquet"
AUDIT_PATH  = CACHE_DIR / "sensors_audit.parquet"


@dataclass
class ConfigSchema:
    sheet: str
    n_rows: int
    n_cols: int
    teros12_blocks: list[tuple[int, int, int]]
    teros21_blocks: list[tuple[int, int]]
    battery_pct: int | None
    battery_mv: int | None
    baro_kpa: int | None
    logger_temp: int | None
    is_standard: bool
    notes: str


def _read_header(filepath: Path, sheet: str) -> pd.DataFrame:
    return pd.read_excel(filepath, sheet_name=sheet, header=None, nrows=3)


def _detect_schema(filepath: Path, sheet: str) -> ConfigSchema | None:
    header = _read_header(filepath, sheet)
    if len(header) < 3:
        return None

    sensor_row = header.iloc[1].astype(str).tolist()
    quant_row  = header.iloc[2].astype(str).str.lower().tolist()
    n_cols = len(sensor_row)
    if n_cols == 0:
        return None

    teros12_blocks: list[tuple[int, int, int]] = []
    teros21_blocks: list[tuple[int, int]] = []
    battery_pct = battery_mv = baro_kpa = logger_temp = None

    i = 1  # col 0 is timestamp
    while i < n_cols:
        sensor = sensor_row[i]
        if sensor in ("nan", "NaN", "None") or pd.isna(header.iloc[1, i]):
            i += 1
            continue
        if "TEROS 12" in sensor:
            if (i + 2 < n_cols
                and "water content" in (quant_row[i] or "")
                and "soil temperature" in (quant_row[i + 1] or "")
                and "bulk ec" in (quant_row[i + 2] or "")):
                teros12_blocks.append((i, i + 1, i + 2))
                i += 3
                continue
            i += 1
            continue
        if "TEROS 21" in sensor:
            if (i + 1 < n_cols
                and "matric potential" in (quant_row[i] or "")
                and "soil temperature" in (quant_row[i + 1] or "")):
                teros21_blocks.append((i, i + 1))
                i += 2
                continue
            i += 1
            continue
        if "Battery" in sensor:
            q = quant_row[i] or ""
            if "battery percent" in q:
                battery_pct = i
            elif "battery voltage" in q:
                battery_mv = i
            i += 1
            continue
        if "Barometer" in sensor:
            q = quant_row[i] or ""
            if "reference pressure" in q:
                baro_kpa = i
            elif "logger temperature" in q:
                logger_temp = i
            i += 1
            continue
        i += 1

    n_t12 = len(teros12_blocks)
    n_t21 = len(teros21_blocks)
    is_standard = (n_t12 in (1, 2, 3)) and (n_t21 in (0, 1, 2, 3)) and (n_t21 <= n_t12)
    notes = [] if is_standard else [f"non-standard: T12={n_t12}, T21={n_t21}"]

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb[sheet]
        n_rows = (ws.max_row or 4) - 3
        wb.close()
    except Exception:
        n_rows = 0

    return ConfigSchema(
        sheet=sheet, n_rows=n_rows, n_cols=n_cols,
        teros12_blocks=teros12_blocks,
        teros21_blocks=teros21_blocks,
        battery_pct=battery_pct, battery_mv=battery_mv,
        baro_kpa=baro_kpa, logger_temp=logger_temp,
        is_standard=is_standard, notes="; ".join(notes),
    )


def _load_config_sheet(filepath: Path, schema: ConfigSchema) -> pd.DataFrame:
    raw = pd.read_excel(filepath, sheet_name=schema.sheet, header=None, skiprows=3)
    if len(raw) == 0:
        return pd.DataFrame()

    out = pd.DataFrame()
    out["timestamp"] = pd.to_datetime(raw[0], errors="coerce")

    for k in range(min(len(schema.teros12_blocks), 3)):
        depth = DEPTH_ORDER[k]
        v, t, e = schema.teros12_blocks[k]
        out[f"vwc_{depth}"]  = pd.to_numeric(raw[v], errors="coerce")
        out[f"temp_{depth}"] = pd.to_numeric(raw[t], errors="coerce")
        out[f"ec_{depth}"]   = pd.to_numeric(raw[e], errors="coerce")

    for k in range(min(len(schema.teros21_blocks), 3)):
        depth = DEPTH_ORDER[k]
        mp, mp_t = schema.teros21_blocks[k]
        out[f"mp_{depth}"]      = pd.to_numeric(raw[mp], errors="coerce")
        out[f"mp_temp_{depth}"] = pd.to_numeric(raw[mp_t], errors="coerce")

    if schema.battery_pct is not None:
        out["battery_pct"] = pd.to_numeric(raw[schema.battery_pct], errors="coerce")
    if schema.battery_mv is not None:
        out["battery_mv"]  = pd.to_numeric(raw[schema.battery_mv], errors="coerce")
    if schema.baro_kpa is not None:
        out["baro_kpa"]    = pd.to_numeric(raw[schema.baro_kpa], errors="coerce")
    if schema.logger_temp is not None:
        out["logger_temp"] = pd.to_numeric(raw[schema.logger_temp], errors="coerce")

    out = out.dropna(subset=["timestamp"])
    out["_source_sheet"] = schema.sheet
    return out


def load_plot_all_configs(
    filepaths: list[Path],
    plot_id: str,
    audit: list[dict] | None = None,
) -> pd.DataFrame:
    frames = []
    for fp in filepaths:
        fp = Path(fp)
        if not fp.exists():
            logger.warning("Plot %s: missing file %s", plot_id, fp)
            continue
        wb = openpyxl.load_workbook(fp, read_only=True)
        sheets = sorted(
            (s for s in wb.sheetnames if s.startswith("Processed Data Config")),
            key=lambda s: int(re.search(r"\d+", s).group()),
        )
        wb.close()

        for sheet in sheets:
            schema = _detect_schema(fp, sheet)
            if schema is None:
                continue
            if not schema.is_standard:
                logger.warning("Plot %s: skipping %s in %s (%s, n_rows~%d)",
                               plot_id, sheet, fp.name, schema.notes, schema.n_rows)
                if audit is not None:
                    audit.append({
                        "plot_id": plot_id, "source_file": fp.name, "sheet": sheet,
                        "n_rows_raw": schema.n_rows, "n_rows_kept": 0,
                        "n_t12": len(schema.teros12_blocks),
                        "n_t21": len(schema.teros21_blocks),
                        "is_standard": False, "kept": False, "notes": schema.notes,
                    })
                continue
            df = _load_config_sheet(fp, schema)
            if df.empty:
                continue
            df["_source_file"] = fp.name
            frames.append(df)
            if audit is not None:
                audit.append({
                    "plot_id": plot_id, "source_file": fp.name, "sheet": sheet,
                    "n_rows_raw": schema.n_rows, "n_rows_kept": len(df),
                    "n_t12": len(schema.teros12_blocks),
                    "n_t21": len(schema.teros21_blocks),
                    "min_ts": df["timestamp"].min(), "max_ts": df["timestamp"].max(),
                    "is_standard": True, "kept": True, "notes": "",
                })

    if not frames:
        logger.error("Plot %s: no usable configs", plot_id)
        return pd.DataFrame()

    combined = pd.concat(frames, ignore_index=True).sort_values("timestamp")
    sensor_cols = [c for c in combined.columns
                   if c.startswith(("vwc_", "temp_", "ec_", "mp_"))]
    combined["_completeness"] = combined[sensor_cols].notna().sum(axis=1)
    combined = (combined.sort_values(["timestamp", "_completeness"])
                          .drop_duplicates("timestamp", keep="last")
                          .drop(columns=["_completeness"]))
    combined["plot_id"]   = plot_id
    combined["treatment"] = TREATMENT_MAP[plot_id]
    combined["depth_map"] = "deep_to_shallow_thermal_validated"
    return combined.reset_index(drop=True)


def load_all_sensors(use_cache: bool = True) -> tuple[pd.DataFrame, pd.DataFrame]:
    if use_cache and CACHE_PATH.exists():
        logger.info("Loading cached sensor data from %s", CACHE_PATH)
        return pd.read_parquet(CACHE_PATH), pd.read_parquet(AUDIT_PATH)

    audit: list[dict] = []
    frames: list[pd.DataFrame] = []
    for plot_id in PLOT_FILES:
        fps = list(PLOT_FILES[plot_id])  # all incremental downloads, chronological
        existing = [fp for fp in fps if Path(fp).exists()]
        if not existing:
            logger.warning("No files for plot %s", plot_id)
            continue
        logger.info("Plot %s: loading from %s", plot_id, [fp.name for fp in existing])
        df = load_plot_all_configs(existing, plot_id, audit=audit)
        if len(df) > 0:
            logger.info("  -> %d rows  [%s -> %s]",
                        len(df), df["timestamp"].min(), df["timestamp"].max())
            frames.append(df)

    sensors = (pd.concat(frames, ignore_index=True)
                  .sort_values(["plot_id", "timestamp"])
                  .reset_index(drop=True))
    for c in ("_source_sheet", "_source_file"):
        if c in sensors.columns:
            sensors = sensors.drop(columns=[c])

    audit_df = pd.DataFrame(audit)

    CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    sensors.to_parquet(CACHE_PATH, index=False)
    audit_df.to_parquet(AUDIT_PATH, index=False)
    logger.info("Cached %d sensor rows to %s", len(sensors), CACHE_PATH)
    logger.info("Cached %d audit rows to %s", len(audit_df), AUDIT_PATH)
    return sensors, audit_df


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s %(levelname)s %(message)s")
    df, audit = load_all_sensors(use_cache=False)
    print(f"\nDataset shape: {df.shape}")
    print(f"Time range: {df['timestamp'].min()} -> {df['timestamp'].max()}")
    print(df.groupby("plot_id")["timestamp"].agg(["min", "max", "count"]).to_string())
